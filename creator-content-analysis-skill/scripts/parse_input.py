#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据解析模块 — 将各种输入格式统一转换为标准 JSON 结构
支持：JSON、CSV、Markdown、TXT、手动粘贴
"""

import json
import csv
import os
import re
from typing import Dict, List, Optional, Union


class InputParser:
    """输入数据解析器，将各种格式转换为标准结构"""

    # 标准字段映射（CSV列名 -> 标准字段名）
    FIELD_MAPPING = {
        "id": "id",
        "title": "title",
        "标题": "title",
        "url": "url",
        "链接": "url",
        "link": "url",
        "publish_time": "publish_time",
        "发布时间": "publish_time",
        "date": "publish_time",
        "description": "description",
        "描述": "description",
        "简介": "description",
        "transcript": "transcript",
        "转写稿": "transcript",
        "文案": "transcript",
        "脚本": "transcript",
        "cover_text": "cover_text",
        "封面文字": "cover_text",
        "封面": "cover_text",
        "hashtags": "hashtags",
        "标签": "hashtags",
        "话题": "hashtags",
        "likes": "likes",
        "点赞": "likes",
        "点赞数": "likes",
        "comments": "comments",
        "评论": "comments",
        "评论数": "comments",
        "shares": "shares",
        "分享": "shares",
        "转发": "shares",
        "转发数": "shares",
        "favorites": "favorites",
        "收藏": "favorites",
        "收藏数": "favorites",
        "duration": "duration",
        "时长": "duration",
        "views": "views",
        "播放量": "views",
        "播放": "views",
    }

    def __init__(self):
        self.data = None

    def load(self, input_source: str, format_hint: Optional[str] = None) -> Dict:
        """
        加载输入数据

        Args:
            input_source: 文件路径或直接的文本内容
            format_hint: 格式提示 (json/csv/markdown/txt)

        Returns:
            标准化的数据字典
        """
        # 判断是文件路径还是直接内容
        if os.path.isfile(input_source):
            return self._load_from_file(input_source, format_hint)
        else:
            # 尝试作为 JSON 字符串解析
            try:
                data = json.loads(input_source)
                self.data = self._normalize_data(data)
                return self.data
            except json.JSONDecodeError:
                pass

            # 尝试作为 CSV 文本解析
            if "," in input_source and "\n" in input_source:
                return self._parse_csv_text(input_source)

            # 尝试作为 Markdown 解析
            if "#" in input_source or "|" in input_source:
                return self._parse_markdown_text(input_source)

            # 作为纯文本处理
            return self._parse_plain_text(input_source)

    def _load_from_file(self, file_path: str, format_hint: Optional[str] = None) -> Dict:
        """从文件加载数据"""
        ext = os.path.splitext(file_path)[1].lower()

        if format_hint:
            ext = f".{format_hint}"

        if ext == ".json":
            return self._load_json(file_path)
        elif ext == ".csv":
            return self._load_csv(file_path)
        elif ext in (".md", ".markdown"):
            return self._load_markdown(file_path)
        elif ext in (".txt", ".text"):
            return self._load_txt(file_path)
        elif ext in (".xlsx", ".xls"):
            return self._load_excel(file_path)
        else:
            # 尝试自动识别
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read(1000)

            if content.strip().startswith("{"):
                return self._load_json(file_path)
            elif "," in content.split("\n")[0]:
                return self._load_csv(file_path)
            else:
                return self._load_txt(file_path)

    def _load_json(self, file_path: str) -> Dict:
        """加载 JSON 文件"""
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.data = self._normalize_data(data)
        return self.data

    def _load_csv(self, file_path: str) -> Dict:
        """加载 CSV 文件"""
        contents = []
        creator_info = {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""}

        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                normalized_row = {}
                for key, value in row.items():
                    # 映射字段名
                    standard_key = self.FIELD_MAPPING.get(key.strip().lower(), key.strip().lower())
                    normalized_row[standard_key] = value

                # 处理标签字段
                if "hashtags" in normalized_row and isinstance(normalized_row["hashtags"], str):
                    normalized_row["hashtags"] = [t.strip() for t in normalized_row["hashtags"].split(",") if t.strip()]

                # 处理数值字段
                for num_field in ["likes", "comments", "shares", "favorites", "views"]:
                    if num_field in normalized_row:
                        try:
                            normalized_row[num_field] = int(normalized_row[num_field].replace(",", "").replace("万", "0000").replace("k", "000").replace("K", "000"))
                        except (ValueError, AttributeError):
                            normalized_row[num_field] = 0

                # 确保有 id
                if "id" not in normalized_row or not normalized_row["id"]:
                    normalized_row["id"] = str(i + 1)

                contents.append(normalized_row)

        self.data = {
            "creator": creator_info,
            "contents": contents
        }
        return self.data

    def _load_markdown(self, file_path: str) -> Dict:
        """加载 Markdown 文件"""
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return self._parse_markdown_text(content)

    def _load_txt(self, file_path: str) -> Dict:
        """加载 TXT 文件"""
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return self._parse_plain_text(content)

    def _load_excel(self, file_path: str) -> Dict:
        """加载 Excel 文件（需要 openpyxl）"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl 才能读取 Excel 文件: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(cell.value)

        # 读取数据
        contents = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row_dict = {}
            for j, value in enumerate(row):
                if j < len(headers) and headers[j]:
                    standard_key = self.FIELD_MAPPING.get(str(headers[j]).strip().lower(), str(headers[j]).strip().lower())
                    row_dict[standard_key] = value if value is not None else ""

            # 处理标签
            if "hashtags" in row_dict and isinstance(row_dict["hashtags"], str):
                row_dict["hashtags"] = [t.strip() for t in row_dict["hashtags"].split(",") if t.strip()]

            # 处理数值
            for num_field in ["likes", "comments", "shares", "favorites", "views"]:
                if num_field in row_dict and row_dict[num_field]:
                    try:
                        row_dict[num_field] = int(float(row_dict[num_field]))
                    except (ValueError, TypeError):
                        row_dict[num_field] = 0

            if "id" not in row_dict or not row_dict["id"]:
                row_dict["id"] = str(i + 1)

            contents.append(row_dict)

        wb.close()

        self.data = {
            "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
            "contents": contents
        }
        return self.data

    def _parse_csv_text(self, text: str) -> Dict:
        """解析 CSV 文本"""
        import io
        reader = csv.DictReader(io.StringIO(text))
        contents = []
        for i, row in enumerate(reader):
            normalized_row = {}
            for key, value in row.items():
                standard_key = self.FIELD_MAPPING.get(key.strip().lower(), key.strip().lower())
                normalized_row[standard_key] = value

            if "hashtags" in normalized_row and isinstance(normalized_row["hashtags"], str):
                normalized_row["hashtags"] = [t.strip() for t in normalized_row["hashtags"].split(",") if t.strip()]

            for num_field in ["likes", "comments", "shares", "favorites", "views"]:
                if num_field in normalized_row:
                    try:
                        normalized_row[num_field] = int(normalized_row[num_field].replace(",", ""))
                    except (ValueError, AttributeError):
                        normalized_row[num_field] = 0

            if "id" not in normalized_row or not normalized_row["id"]:
                normalized_row["id"] = str(i + 1)

            contents.append(normalized_row)

        self.data = {
            "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
            "contents": contents
        }
        return self.data

    def _parse_markdown_text(self, text: str) -> Dict:
        """解析 Markdown 文本（表格格式）"""
        contents = []
        lines = text.strip().split("\n")

        # 查找表格
        table_lines = []
        in_table = False
        headers = []

        for line in lines:
            line = line.strip()
            if "|" in line:
                if not in_table:
                    # 第一个表格行作为表头
                    headers = [h.strip() for h in line.split("|") if h.strip()]
                    in_table = True
                elif "---" not in line:
                    # 数据行
                    cells = [c.strip() for c in line.split("|") if c.strip()]
                    if len(cells) == len(headers):
                        row = {}
                        for j, cell in enumerate(cells):
                            standard_key = self.FIELD_MAPPING.get(headers[j].lower(), headers[j].lower())
                            row[standard_key] = cell
                        contents.append(row)

        # 处理标签和数值
        for i, row in enumerate(contents):
            if "hashtags" in row and isinstance(row["hashtags"], str):
                row["hashtags"] = [t.strip() for t in row["hashtags"].split(",") if t.strip()]
            for num_field in ["likes", "comments", "shares", "favorites", "views"]:
                if num_field in row:
                    try:
                        row[num_field] = int(row[num_field].replace(",", "").replace("万", "0000"))
                    except (ValueError, AttributeError):
                        row[num_field] = 0
            if "id" not in row or not row["id"]:
                row["id"] = str(i + 1)

        self.data = {
            "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
            "contents": contents
        }
        return self.data

    def _parse_plain_text(self, text: str) -> Dict:
        """解析纯文本（尝试提取标题和文案）"""
        contents = []
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]

        for i, para in enumerate(paragraphs):
            lines = para.split("\n")
            title = lines[0].strip() if lines else ""
            transcript = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""

            contents.append({
                "id": str(i + 1),
                "title": title,
                "transcript": transcript,
                "url": "",
                "publish_time": "",
                "description": "",
                "cover_text": "",
                "hashtags": [],
                "likes": 0,
                "comments": 0,
                "shares": 0,
                "favorites": 0,
                "duration": "",
                "views": 0
            })

        self.data = {
            "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
            "contents": contents
        }
        return self.data

    def _normalize_data(self, data: Dict) -> Dict:
        """标准化数据结构"""
        # 如果已经有标准结构
        if "creator" in data and "contents" in data:
            # 确保 creator 字段完整
            default_creator = {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""}
            for key in default_creator:
                if key not in data["creator"]:
                    data["creator"][key] = default_creator[key]

            # 确保每条内容字段完整
            default_content = {
                "id": "", "title": "", "url": "", "publish_time": "",
                "description": "", "transcript": "", "cover_text": "",
                "hashtags": [], "likes": 0, "comments": 0, "shares": 0,
                "favorites": 0, "duration": "", "views": 0
            }
            for i, content in enumerate(data["contents"]):
                for key in default_content:
                    if key not in content:
                        content[key] = default_content[key]
                if not content.get("id"):
                    content["id"] = str(i + 1)

            return data

        # 如果只有 contents（列表形式）
        if isinstance(data, list):
            return {
                "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
                "contents": data
            }

        # 其他情况，尝试推断
        return {
            "creator": {"name": "", "platform": "", "profile_url": "", "bio": "", "followers": "", "following": "", "total_posts": ""},
            "contents": []
        }

    def get_stats(self) -> Dict:
        """获取数据统计信息"""
        if not self.data:
            return {"loaded": False}

        contents = self.data.get("contents", [])
        total_likes = sum(c.get("likes", 0) for c in contents)
        total_comments = sum(c.get("comments", 0) for c in contents)
        total_shares = sum(c.get("shares", 0) for c in contents)
        total_favorites = sum(c.get("favorites", 0) for c in contents)
        total_views = sum(c.get("views", 0) for c in contents)

        return {
            "loaded": True,
            "total_contents": len(contents),
            "total_likes": total_likes,
            "total_comments": total_comments,
            "total_shares": total_shares,
            "total_favorites": total_favorites,
            "total_views": total_views,
            "avg_likes": total_likes / len(contents) if contents else 0,
            "avg_comments": total_comments / len(contents) if contents else 0,
            "has_transcript": sum(1 for c in contents if c.get("transcript")),
            "has_hashtags": sum(1 for c in contents if c.get("hashtags")),
        }


def parse_input(input_source: str, format_hint: Optional[str] = None) -> Dict:
    """
    便捷函数：解析输入数据

    Args:
        input_source: 文件路径或文本内容
        format_hint: 格式提示

    Returns:
        标准化的数据字典
    """
    parser = InputParser()
    return parser.load(input_source, format_hint)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        data = parse_input(input_file)
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        print("用法: python parse_input.py <输入文件路径>")
        print("支持格式: JSON, CSV, Markdown, TXT, Excel")
